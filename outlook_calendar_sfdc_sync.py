from tkinter import *
import tkinter as tk
import os
import datetime as dt
from dateutil.relativedelta import relativedelta
import pytz
from openpyxl import load_workbook
import win32com.client
import win32api
import xlrd
from simple_salesforce import Salesforce
import pandas as pd
from openpyxl.styles import Font, Color, PatternFill
import win32timezone

xl = win32com.client.Dispatch("Excel.Application")
datenow = dt.datetime.today()
lastmonth = datenow - relativedelta(months=1)
root = tk.Tk()
root.title("Outlook-SFDC Sync")

# Add a grid
mainframe = Frame(root)
mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
mainframe.columnconfigure(0, weight=1)
mainframe.rowconfigure(0, weight=1)
mainframe.pack(pady=25, padx=25)

# Create a Tkinter variable
tkvar = StringVar(root)
tkvar2 = StringVar(root)

# Dictionary with options
choices = {'Outlook Download', 'Outlook SFDC Upload',}
tkvar.set('Outlook Download')  # set the default option

popupMenu = OptionMenu(mainframe, tkvar, *choices)
Label(mainframe, text="Please choose which script you want to run:").grid(row=1, column=1)
popupMenu.grid(row=2, column=1)

def ok():
    script = tkvar.get()
    if script == 'Outlook Download':
        create_window()
    elif script ==  'Outlook SFDC Upload':
        root.withdraw()
        wb = xlrd.open_workbook('C:\\SFDC Outlook Synchronization\\SFDC_Admin\\SFDC_Admin.xlsx')
        first_sheet = wb.sheet_by_name("Sheet1")
        a1 = first_sheet.cell(0, 1).value
        a2 = first_sheet.cell(1, 1).value
        a3 = first_sheet.cell(2, 1).value

        try:
            sf = Salesforce(username=a1, password=a2, security_token=a3)
        except Exception:
            win32api.MessageBox(0,
                                "The script cannot run. You need to either 1) Update your Salesforce password (in cell B2) in the following file: C:\SFDC Outlook Synchronization\SFDC_Admin\SFDC_Admin.xlsx and save the file or 2) Check your Internet connectivity.",
                                "Error!",
                                0x00001000)
            root.destroy()
            quit()

        for wbb in xl.Workbooks:
            if wbb.Name == 'Outlook Sync.xlsx':
                wbb.Close(True)

        wb = load_workbook(filename="C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx", read_only=False,keep_vba=False)
        ws = wb.get_sheet_by_name('Outlook Sync')
        file = pd.ExcelFile('C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx')
        df = file.parse('Outlook Sync')
        #df.sort_values(['Upload Event to SFDC?'], ascending=[False], inplace=True)

        g = 1

        for index, row in df.iterrows():
            try:
                if df.iloc[index]['Upload Event to SFDC?'] == "No" or pd.isnull(df.iloc[index]['Upload Event to SFDC?']):
                    g = g + 1
                    continue
                if "Yes" in ws.cell(row=g + 1, column=17).value:
                    g = g + 1
                    continue
            except(Exception):
                pass

            try:
                g = g + 1

                if "'" not in df.iloc[index]['SFDC Contact']:
                    contact = "'" + df.iloc[index]['SFDC Contact'] + "'"
                else:
                    contact = df.iloc[index]['SFDC Contact']

                query_result2 = sf.query_all("SELECT Id FROM Contact Where Email = %s" % contact)
                records2 = query_result2['records']
                df2 = pd.DataFrame(records2)
                df2.drop('attributes', inplace=True, axis=1)

                if "'" not in df.iloc[index]['Assigned To']:
                    owner = "'" + df.iloc[index]['Assigned To'] + "'"
                else:
                    owner = df.iloc[index]['Assigned To']

                query_result4 = sf.query_all("SELECT Id FROM User Where Email = %s" % owner)
                records4 = query_result4['records']
                df4 = pd.DataFrame(records4)
                df4.drop('attributes', inplace=True, axis=1)

                if pd.isnull(df.iloc[index]['Location']):
                    location = ''
                else:
                    location = df.iloc[index]['Location']
                if pd.isnull(df.iloc[index]['Appointment Body']):
                    description = ''
                else:
                    description =  df.iloc[index]['Appointment Body']
                if pd.isnull(df.iloc[index]['SFDC Opportunity']):
                    oppty = ''
                else:
                    oppty = "'" + df.iloc[index]['SFDC Opportunity'] + "'"
                    query_result3 = sf.query_all("SELECT Id FROM Opportunity Where Name = %s" % oppty)
                    records3 = query_result3['records']
                    df3 = pd.DataFrame(records3)
                    df3.drop('attributes', inplace=True, axis=1)
                    oppty = df3.iloc[0]['Id']

                if pd.isnull(df.iloc[index]['Additional Participant #1']):
                    additionalp1 = ""
                    additionalp1id = ""
                else:
                    additionalp1 = df.iloc[index]['Additional Participant #1']
                    additionalp1sfdc = "'" + additionalp1 + "'"
                    try:
                        query_result5 = sf.query_all("SELECT Id FROM User Where Name = %s" % additionalp1sfdc)
                        records5 = query_result5['records']
                        df5 = pd.DataFrame(records5)
                        df5.drop('attributes', inplace=True, axis=1)
                        additionalp1id = df5.iloc[0]['Id']
                    except(Exception):
                        additionalp1id = ""

                if pd.isnull(df.iloc[index]['Additional Participant #2']):
                    additionalp2 = ""
                    additionalp2id = ""
                else:
                    additionalp2 = df.iloc[index]['Additional Participant #2']
                    additionalp2sfdc = "'" + additionalp2 + "'"
                    try:
                        query_result6 = sf.query_all("SELECT Id FROM User Where Name = %s" % additionalp2sfdc)
                        records6 = query_result6['records']
                        df6 = pd.DataFrame(records6)
                        df6.drop('attributes', inplace=True, axis=1)
                        additionalp2id = df6.iloc[0]['Id']
                    except(Exception):
                        additionalp2id = ""
                if pd.isnull(df.iloc[index]['Additional Participant #3']):
                    additionalp3 = ""
                    additionalp3id = ""
                else:
                    additionalp3 = df.iloc[index]['Additional Participant #3']
                    additionalp3sfdc = "'" + additionalp3 + "'"
                    try:
                        query_result7 = sf.query_all("SELECT Id FROM User Where Name = %s" % additionalp3sfdc)
                        records7 = query_result7['records']
                        df7 = pd.DataFrame(records7)
                        df7.drop('attributes', inplace=True, axis=1)
                        additionalp3id = df7.iloc[0]['Id']
                    except(Exception):
                        additionalp3id = ""

                local = pytz.timezone("Australia/Sydney")

                startdate = dt.datetime.strptime(df.iloc[index]['Start'], "%Y-%B-%d %I:%M%p")
                local_dt = local.localize(startdate, is_dst=None)
                startdate = local_dt.astimezone(pytz.utc)
                startdate = startdate.strftime("%Y-%m-%dT%H:%M:%S")

                enddate = dt.datetime.strptime(df.iloc[index]['End'], "%Y-%B-%d %I:%M%p")
                local_dt = local.localize(enddate, is_dst=None)
                enddate = local_dt.astimezone(pytz.utc)
                enddate = enddate.strftime("%Y-%m-%dT%H:%M:%S")

                if df.iloc[index]['Event Type'] == "Customer.Value.Assurance":
                    cv = 'Customer Value Assurance'
                else:
                    cv = df.iloc[index]['Event Type']

                if cv == "Customer Value Assurance" and (df.iloc[index]['Event Sub-Type'] != "Onboarding Conversation" and df.iloc[index]['Event Sub-Type'] != "Service Orientation" and df.iloc[index]['Event Sub-Type'] != "Mid-Year Check In" and df.iloc[index]['Event Sub-Type'] != "Relationship Building" and df.iloc[index]['Event Sub-Type'] != "Recruiting"):
                    ws.cell(row=g,column=17).value = "No - your event type and event sub-type do not align. Please fix this and try to upload again."
                    ft = Font()
                    ft.underline = 'none'  # add single underline
                    ft.color = Color(rgb='FF000000')  # add blue color
                    ws.cell(row=g, column=17).font = ft
                    wb.save("C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx")
                    continue
                if cv == "Commercial" and (df.iloc[index]['Event Sub-Type'] != "Prospecting" and df.iloc[index]['Event Sub-Type'] != "Sales Introduction" and df.iloc[index]['Event Sub-Type'] != "Sales Follow Up" and df.iloc[index]['Event Sub-Type'] != "Closing Call"):
                    ws.cell(row=g,column=17).value = "No - your event type and event sub-type do not align. Please fix this and try to upload again."
                    ft = Font()
                    ft.underline = 'none'  # add single underline
                    ft.color = Color(rgb='FF000000')  # add blue color
                    ws.cell(row=g, column=17).font = ft
                    wb.save("C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx")
                    continue
                if cv == "Consultation" and (df.iloc[index]['Event Sub-Type'] != "Interview" and df.iloc[index]['Event Sub-Type'] != "Prewire/Scoping" and df.iloc[index]['Event Sub-Type'] != "Advisory/Content Review" and df.iloc[index]['Event Sub-Type'] != "Results Delivery"):
                    ws.cell(row=g,column=17).value = "No - your event type and event sub-type do not align. Please fix this and try to upload again."
                    ft = Font()
                    ft.underline = 'none'  # add single underline
                    ft.color = Color(rgb='FF000000')  # add blue color
                    ws.cell(row=g, column=17).font = ft
                    wb.save("C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx")
                    continue

                results = sf.Event.create({'OwnerId': df4.iloc[0]['Id'], 'WhoId': df2.iloc[0]['Id'], 'WhatId': oppty, 'Event_Type__c': cv, 'Event_Sub_Type__c': df.iloc[index]['Event Sub-Type'],
                     'Channel__c': df.iloc[index]['Channel'], 'Status__c': df.iloc[index]['Event Status'], 'Subject': df.iloc[index]['Subject'],
                     'ShowAs': 'Busy','StartDateTime': startdate, 'EndDateTime': enddate, 'Location': location, 'Description': description,
                     'Additional_Participant_1_Login__c': additionalp1, 'Additional_Participant_1_Id__c': additionalp1id,
                     'Additional_Participant_2_Login__c': additionalp2, 'Additional_Participant_2_Id__c': additionalp2id,
                     'Additional_Participant_3_Login__c': additionalp3, 'Additional_Participant_3_Id__c': additionalp3id, 'Maps_Category__c': df.iloc[index]['Outlook Id']})

                link = "https://na29.salesforce.com/" + results['id']

                ws.cell(row=g, column=17).value = '=HYPERLINK("{}", "{}")'.format(link, "Yes")
                ws.cell(row=g, column=19).value = '=HYPERLINK("{}", "{}")'.format(link, "Yes - has been just uploaded or previously uploaded")
                ft = Font()
                ft.underline = 'single'  # add single underline
                ft.color = Color(rgb='000000FF')  # add blue color
                ws.cell(row=g, column=17).font = ft
                ws.cell(row=g, column=19).font = ft

                FILL_Green = PatternFill(start_color='00FF00',
                                        end_color='00FF00',
                                        fill_type='solid')

                for row in ws['A'+ str(g) +":"+"S"+ str(g)]:
                    for cell in row:
                        cell.fill = FILL_Green

                wb.save("C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx")
                continue

            except(Exception):
                ws.cell(row=g, column=17).value = "No - either the contact/opportunity doesn't exist in SFDC or you left a key field such as 'Event Type' blank or incorrectly spelled. Please make edits and try uploading again."
                ft = Font()
                ft.underline = 'none'  # add single underline
                ft.color = Color(rgb='FF000000')  # add blue color
                ws.cell(row=g, column=17).font = ft
                wb.save("C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx")
                continue

        wb15 = load_workbook(filename='C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx')
        wb15.save("C:\\SFDC Outlook Synchronization\\Previous Uploads\\Outlook Sync - " + datenow.strftime("%d %B %Y %I-%M %p") + ".xlsx")
        wb2 = xl.Workbooks.Open('C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx')
        xl.Visible = True
        win32api.MessageBox(0,
                            "The Outlook-SFDC upload process is complete! Please check the events you created by clicking the hyperlinks under the column 'Upload Sucessful?'",
                            "Outlook-SFDC Upload Process Complete!", 0x00001000)
        root.destroy()


def rangeok(e1, e2, e3, e4, e5, e6, name, window):
    f1 = int(e1.get())
    f2 = int(e2.get())
    f3 = int(e3.get())
    f4 = int(e4.get())
    f5 = int(e5.get())
    f6 = int(e6.get())
    internal = tkvar2.get()
    shared = str(name.get())

    if 'withdrawn' == root.state():
        try:
            begin = dt.datetime(f3, f2, f1, 0, 0)
            begin = begin.strftime("%Y-%m-%d %H:%M")
            end = dt.datetime(f6, f5, f4, 23, 59)
            end = end.strftime("%Y-%m-%d %H:%M")
            if begin <= end:
                download(begin, end, internal, shared, window)
            else:
                win32api.MessageBox(0,
                                    "The start date is greater than the end date. Please fix this and click 'Ok' again.",
                                    "Error!",
                                    0x00001000)
        except(ValueError):
            win32api.MessageBox(0,
                                "The date(s) you inputted are not real date(s). Please input real date(s) and click 'Ok' again. =)",
                                "Error!",
                                0x00001000)

def create_window():
    window = tk.Toplevel(root)
    window.geometry('%dx%d+%d+%d' % (355, 315, x, y))
    Label(window, text="Select from the drop-down what you want to download:").place(x=0, y=0, width=350)
    choices = {'Today', 'Past Week', 'Past Month', 'Next Week', 'Next Month'}
    tkvar.set('Today')  # set the default option

    popupMenu = OptionMenu(window, tkvar, *choices)
    popupMenu.place(x=70, y=25, width=100)

    Label(window, text="OR", font = 'bold', foreground ='red').place(x=155, y=80, width=50)
    Label(window, text="Input the start/end dates of what you want to download:").place(x=0, y=125, width=350)
    Label(window, text="Start Date").place(x=5, y=150, width=50)
    Label(window, text="End Date").place(x=180, y=150, width=50)
    e1 = Entry(window, justify='center')
    e2 = Entry(window, justify='center')
    e3 = Entry(window, justify='center')
    e4 = Entry(window, justify='center')
    e5 = Entry(window, justify='center')
    e6 = Entry(window, justify='center')
    e1.place(x=60, y=150, width=25)
    Label(window, text="DD").place(x=60, y=170, width=25)
    e1.insert(0, lastmonth.day)
    e2.place(x=90, y=150, width=25)
    Label(window, text="MM").place(x=90, y=170, width=25)
    e2.insert(0, lastmonth.month)
    e3.place(x=120, y=150, width=40)
    Label(window, text="YYYY").place(x=120, y=170, width=40)
    e3.insert(0, lastmonth.year)
    e4.place(x=240, y=150, width=25)
    Label(window, text="DD").place(x=240, y=170, width=25)
    e4.insert(0, datenow.day)
    e5.place(x=270, y=150, width=25)
    Label(window, text="MM").place(x=270, y=170, width=25)
    e5.insert(0, datenow.month)
    e6.place(x=300, y=150, width=45)
    Label(window, text="YYYY").place(x=300, y=170, width=40)
    e6.insert(0, datenow.year)

    rangeb = Button(window, text="Ok", bg='pale green', command=lambda: rangeok(e1, e2, e3, e4, e5, e6, name, window))
    rangeb.place(x=165, y=185, width=35)

    Label(window, text="-------------------------------Optional-------------------------------").place(x=0, y=235)

    Label(window, text="Include Internal Planners?").place(x=1, y=265, width=150)
    choices2 = {'Yes', 'No'}
    tkvar2.set('No')  # set the default option
    popupMenu2 = OptionMenu(window, tkvar2, *choices2)
    popupMenu2.place(x=50, y=282, height = 27, width=50)

    Label(window, text="Download a Shared Calendar?").place(x=185, y=265, width=160)
    Label(window, text="Write in Name").place(x=170, y=287, width=80)
    name = Entry(window)
    name.place(x=255, y=287, width=95)
    name.insert(0, "")

    choiceb = Button(window, text="Ok", bg='pale green', command=lambda: choiceok(name, window))
    choiceb.place(x=225, y=28, width=35)

    root.withdraw()

    window.protocol('WM_DELETE_WINDOW', exit)

def exit():
    root.destroy()

def download(begin, end, internal, shared, window):
    window.withdraw()
    wb = xlrd.open_workbook('C:\\SFDC Outlook Synchronization\\SFDC_Admin\\SFDC_Admin.xlsx')
    first_sheet = wb.sheet_by_name("Sheet1")
    a1 = first_sheet.cell(0, 1).value
    a2 = first_sheet.cell(1, 1).value
    a3 = first_sheet.cell(2, 1).value

    try:
        sf = Salesforce(username=a1, password=a2, security_token=a3)
    except Exception:
        win32api.MessageBox(0,
                            "The script cannot run. You need to either 1) Update your Salesforce password (in cell B2) in the following file: C:\SFDC Outlook Synchronization\SFDC_Admin\SFDC_Admin.xlsx and save the file or 2) Check your Internet connectivity.",
                            "Error!",
                            0x00001000)
        root.destroy()
        quit()

    for wbb in xl.Workbooks:
        if wbb.Name == 'Outlook Sync.xlsx':
            wbb.Close(True)

    wb = load_workbook(filename="C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx", read_only=False, keep_vba=False)
    ws = wb.get_sheet_by_name('Outlook Sync')

    FILL_NONE = PatternFill(start_color='FFFFFF',
                          end_color='FFFFFF',
                          fill_type='none')

    for row in ws['A2:S1000']:
        for cell in row:
            cell.value = None
            cell.fill = FILL_NONE

    wb.save("C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx")

    if shared == "":
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        calendar = outlook.GetDefaultFolder(9)
        appts = calendar.Items
    else:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        myRecipient = outlook.CreateRecipient(shared)
        myRecipient.Resolve()
        if myRecipient.Resolved == True:
            try:
                othercalendar = outlook.GetSharedDefaultFolder(myRecipient, 9)
                appts = othercalendar.Items
            except(Exception):
                win32api.MessageBox(0,
                                    "You don't have access to the shared calendar you inputted. Please ask the person for access to their calendar and try again.",
                                    "Error",
                                    0x00001000)
                root.destroy()
                quit()

        else:
            win32api.MessageBox(0,
                                "The name you inputted for their shared calendar does not exist in the corporate directory. Please choose a different name.",
                                "Error",
                                0x00001000)
            root.destroy()
            quit()

    datenow = dt.datetime.today()

    try:
        appts.Sort("[Start]")
        appts.IncludeRecurrences = True
        appts = appts.Restrict("[Start] >= '" + begin + "' AND [Start] <= '" + end + "'")
    except(Exception):
        win32api.MessageBox(0,
                            "An error occured trying to download the shared calendar. Please ensure you have the correct permissions and try again.",
                            "Error",
                            0x00001000)
        root.destroy()
        quit()

    local = pytz.timezone("Australia/Sydney")
    local_dt = local.localize(datenow, is_dst=None)
    datenow = local_dt.astimezone(pytz.utc)

    g = 1

    for appt in appts:
        attlist = []
        otherlist = []
        opptylist = []
        confirmedlist = []
        finallist = []
        ceblist = []
        wodkalist = []
        attendee = appt.RequiredAttendees
        attlist = attendee.split("; ")

        for att in attlist:
            search_string = att
            recipient = outlook.Session.CreateRecipient(search_string)
            recipient.Resolve()
            if recipient.Resolved == True:
                ae = recipient.AddressEntry
                email_address = None
                try:
                    if 'EX' == ae.Type:
                        eu = ae.GetExchangeUser()
                        email_address = eu.PrimarySmtpAddress

                    if 'SMTP' == ae.Type:
                        email_address = ae.Address

                    if "cebglobal.com" not in email_address and "gartner.com" not in email_address and "evanta.com" not in email_address and "executiveboard.com" not in email_address:
                        otherlist.append(email_address)
                    else:
                        if email_address == a1 and shared == '':
                            continue
                        attsplit = att.split(", ")
                        attsplit = attsplit[1] + " " + attsplit[0]
                        if attsplit == shared and shared != '':
                            continue
                        ceblist.append(attsplit)

                except(Exception):
                    pass

        if otherlist != []:
            for attd in otherlist:
                if confirmedlist != []:
                    break
                if "'" not in attd:
                    contact = "'" + attd + "'"
                else:
                    contact = attd
                query_result1 = sf.query_all("SELECT Id, (SELECT OpportunityId From OpportunityContactRoles) FROM Contact Where Email = %s" % contact)
                records1 = query_result1['records']
                df1 = pd.DataFrame(records1)
                if df1.empty:
                    continue
                else:
                    opptylist.append(attd)
                    df1.drop('attributes', inplace=True, axis=1)
                    for index, row in df1.iterrows():
                        name = (row['OpportunityContactRoles'])
                        if name == None:
                            finallist.append(attd)
                            break
                        else:
                            confirmedlist.append(attd)
                            g = g + 1
                        df1.loc[index, 'OpportunityContactRoles'] = name['records'][0]['OpportunityId']
                        ws.cell(row=g, column=6).value = attd

                        oppty = "'" + df1.iloc[0]['OpportunityContactRoles'] + "'"
                        query_result2 = sf.query_all("SELECT Name FROM Opportunity Where Id = %s" % oppty)
                        records2 = query_result2['records']
                        df2 = pd.DataFrame(records2)
                        ws.cell(row=g, column=7).value = df2.iloc[0]['Name']
                        break
            if confirmedlist == [] and finallist !=[]:
                for attd in finallist:
                    if "'" not in attd:
                        othercontact = "'" + attd + "'"
                    else:
                        othercontact = attd
                    query_result3 = sf.query_all("SELECT Opportunity_Name__c FROM OCCR__c Where Contact_Email__c = %s" % othercontact)
                    records3 = query_result3['records']
                    df3 = pd.DataFrame(records3)
                    if df3.empty:
                        continue
                    else:
                        wodkalist.append(attd)
                        df3.drop('attributes', inplace=True, axis=1)
                        g = g + 1
                        ws.cell(row=g, column=6).value = attd
                        ws.cell(row=g, column=7).value = df3.iloc[0]['Opportunity_Name__c']
                        break

            elif confirmedlist == [] and finallist ==[]:
                g = g + 1
                ws.cell(row=g, column=6).value = otherlist[0]

        elif otherlist == [] and internal == "Yes":
            g = g + 1
        elif otherlist == [] and internal == "No":
            continue

        if confirmedlist == [] and finallist !=[] and wodkalist == []:
            g = g + 1
            ws.cell(row=g, column=6).value = finallist[0]

        ws.cell(row=g, column=1).value = appt.Subject
        ws.cell(row=g, column=2).value = appt.Start.strftime("%Y-%B-%d %I:%M%p")
        ws.cell(row=g, column=3).value = (appt.Start + relativedelta(minutes=appt.Duration)).strftime("%Y-%B-%d %I:%M%p")
        ws.cell(row=g, column=4).value = appt.Location
        ws.cell(row=g, column=5).value = appt.Body
        if shared == '':
            ws.cell(row=g, column=8).value = a1
        else:
            recipient = outlook.Session.CreateRecipient(shared)
            recipient.Resolve()
            if recipient.Resolved == True:
                ae = recipient.AddressEntry
                if 'EX' == ae.Type:
                    eu = ae.GetExchangeUser()
                    email_address = eu.PrimarySmtpAddress

                if 'SMTP' == ae.Type:
                    email_address = ae.Address

                ws.cell(row=g, column=8).value = email_address

        try:
            additionalp1 = ceblist[0]
        except(Exception):
            additionalp1 = ""
        try:
            additionalp2 = ceblist[1]
        except(Exception):
            additionalp2 = ""
        try:
            additionalp3 = ceblist[2]
        except(Exception):
            additionalp3 = ""

        ws.cell(row=g, column=9).value = additionalp1
        ws.cell(row=g, column=10).value = additionalp2
        ws.cell(row=g, column=11).value = additionalp3
        if appt.IsRecurring == True:
            ws.cell(row=g, column=18).value = appt.GlobalAppointmentID + " - " + appt.Start.strftime("%Y-%B-%d %I:%M%p")
        else:
            ws.cell(row=g, column=18).value = appt.GlobalAppointmentID

        try:
            uniqueid = "'" + appt.GlobalAppointmentID + " - " + appt.Start.strftime("%Y-%B-%d %I:%M%p") + "'"
            superowner = "'" + ws.cell(row=g, column=8).value + "'"
            query_result11 = sf.query_all("SELECT Id FROM User Where Email = %s" % superowner)
            records11 = query_result11['records']
            df11 = pd.DataFrame(records11)
            df11.drop('attributes', inplace=True, axis=1)
            superowner = df11.iloc[0]['Id']
            superowner = "'" + superowner + "'"

            query_result10 = sf.query_all("SELECT Id FROM Event WHERE OwnerId = %s AND Maps_Category__c	= %s" % (superowner, uniqueid))
            records10 = query_result10['records']
            df10 = pd.DataFrame(records10)
            if not df10.empty:
                link = "https://na29.salesforce.com/" + df10.iloc[0]['Id']
                ws.cell(row=g, column=19).value = '=HYPERLINK("{}", "{}")'.format(link, "High - Click Here to See the Event in SFDC")
                ft = Font()
                ft.underline = 'single'  # add single underline
                ft.color = Color(rgb='000000FF')  # add blue color
                ws.cell(row=g, column=19).font = ft
            else:
                ws.cell(row=g, column=19).value = "Cannot Determine"
                ft = Font()
                ft.underline = 'none'  # add single underline
                ft.color = Color(rgb='FF000000')  # add blue color
                ws.cell(row=g, column=19).font = ft

                startdate = dt.datetime.strptime(ws.cell(row=g, column=2).value, "%Y-%B-%d %I:%M%p")
                local_dt = local.localize(startdate, is_dst=None)
                startdate = local_dt.astimezone(pytz.utc)
                starthigh = startdate + relativedelta(days=1)
                startlow = startdate - relativedelta(days=1)
                starthigh = starthigh.strftime("%Y-%m-%dT%H:%M:%SZ")
                startlow = startlow.strftime("%Y-%m-%dT%H:%M:%SZ")
                startdate = startdate.strftime("%Y-%m-%dT%H:%M:%SZ")

                query_result16 = sf.query_all("SELECT Id FROM Event WHERE OwnerId = %s AND StartDateTime <= %s AND StartDateTime >= %s" % (superowner, starthigh, startlow))
                records16 = query_result16['records']
                df16 = pd.DataFrame(records16)
                if not df16.empty:
                    link = "https://na29.salesforce.com/" + df16.iloc[0]['Id']
                    ws.cell(row=g, column=19).value = '=HYPERLINK("{}", "{}")'.format(link,"Medium - Click Here to See the Likely Event in SFDC")
                    ft = Font()
                    ft.underline = 'single'  # add single underline
                    ft.color = Color(rgb='000000FF')  # add blue color
                    ws.cell(row=g, column=19).font = ft
                else:
                    ws.cell(row=g, column=19).value = "Cannot Determine"
                    ft = Font()
                    ft.underline = 'none'  # add single underline
                    ft.color = Color(rgb='FF000000')  # add blue color
                    ws.cell(row=g, column=19).font = ft

                    if "'" in ws.cell(row=g, column=6).value:
                        contactemail = ws.cell(row=g, column=6).value
                    else:
                        contactemail = "'" + ws.cell(row=g, column=6).value + "'"

                    query_result18 = sf.query_all("SELECT Id FROM Contact WHERE Email = %s" % (contactemail))
                    records18 = query_result18['records']
                    df18 = pd.DataFrame(records18)
                    if not df18.empty:
                        contactid = "'" + df18.iloc[0]['Id'] + "'"
                        query_result17 = sf.query_all("SELECT Id FROM Event WHERE WhoId = %s AND StartDateTime <= %s AND StartDateTime >= %s" % (contactid, starthigh, startlow))
                        records17 = query_result17['records']
                        df17 = pd.DataFrame(records17)
                        if not df17.empty:
                            link = "https://na29.salesforce.com/" + df17.iloc[0]['Id']
                            ws.cell(row=g, column=19).value = '=HYPERLINK("{}", "{}")'.format(link, "Medium - Click Here to See the Likely Event in SFDC")
                            ft = Font()
                            ft.underline = 'single'  # add single underline
                            ft.color = Color(rgb='000000FF')  # add blue color
                            ws.cell(row=g, column=19).font = ft
                        else:
                            ws.cell(row=g, column=19).value = "Cannot Determine"
                            ft = Font()
                            ft.underline = 'none'  # add single underline
                            ft.color = Color(rgb='FF000000')  # add blue color
                            ws.cell(row=g, column=19).font = ft
                    else:
                        ws.cell(row=g, column=19).value = "Cannot Determine"
                        ft = Font()
                        ft.underline = 'none'  # add single underline
                        ft.color = Color(rgb='FF000000')  # add blue color
                        ws.cell(row=g, column=19).font = ft

        except(Exception):
            ws.cell(row=g, column=19).value = "Cannot Determine"
            ft = Font()
            ft.underline = 'none'  # add single underline
            ft.color = Color(rgb='FF000000')  # add blue color
            ws.cell(row=g, column=19).font = ft

        localy = pytz.timezone("Australia/Sydney")

        datenowy = dt.datetime.today()
        local_dty = localy.localize(datenowy, is_dst=None)
        datenowy = local_dty.astimezone(pytz.utc)

        newstartey = (appt.Start + relativedelta(minutes=appt.Duration)).replace(tzinfo=None)
        local_dty = localy.localize(newstartey, is_dst=None)
        newstartey = local_dty.astimezone(pytz.utc)

        if datenowy >= newstartey:
            ws.cell(row=g, column=12).value = "Completed"
        else:
            ws.cell(row=g, column=12).value = "Scheduled"

        ws.cell(row=g, column=13).value = "Phone/Virtual"

        redFill = PatternFill(start_color='FF8080',
                              end_color='FF8080',
                              fill_type='solid')

        ws.cell(row=g, column=14).fill = redFill
        ws.cell(row=g, column=15).fill = redFill
        ws.cell(row=g, column=16).fill = redFill

        wb.save("C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx")

    wb2 = xl.Workbooks.Open('C:\\SFDC Outlook Synchronization\\Outlook Sync.xlsx')
    xl.Visible = True
    win32api.MessageBox(0, "The Outlook calendar download process is complete! Please check the Excel spreadsheet, make any edits, and mark any events you want to upload to SFDC.",
                        "Outlook Calendar Download Process is Complete!", 0x00001000)

    root.destroy()

def choiceok(name, window):
    dates = tkvar.get()
    begin = dt.datetime.today()
    begin = begin.strftime("%Y-%m-%d %H:%M")
    end = dt.datetime.today().strftime("%Y-%m-%d %H:%M")
    internal = tkvar2.get()
    shared = str(name.get())

    if dates == "Today":
        begin = dt.datetime(datenow.year, datenow.month, datenow.day, 0, 0)
        begin = begin.strftime("%Y-%m-%d %H:%M")
        end = dt.datetime(datenow.year, datenow.month, datenow.day, 23, 59)
        end = end.strftime("%Y-%m-%d %H:%M")
    elif dates == "Past Week":
        begin = dt.datetime(datenow.year, datenow.month, datenow.day, 0, 0)
        begin = begin - relativedelta(days=7)
        begin = begin.strftime("%Y-%m-%d %H:%M")
        end = dt.datetime(datenow.year, datenow.month, datenow.day, 23, 59)
        end = end.strftime("%Y-%m-%d %H:%M")
    elif dates == "Past Month":
        begin = dt.datetime(datenow.year, datenow.month, datenow.day, 0, 0)
        begin = begin - relativedelta(months = 1)
        begin = begin.strftime("%Y-%m-%d %H:%M")
        end = dt.datetime(datenow.year, datenow.month, datenow.day, 23, 59)
        end = end.strftime("%Y-%m-%d %H:%M")
    elif dates == "Next Week":
        begin = dt.datetime(datenow.year, datenow.month, datenow.day, 0, 0)
        begin = begin.strftime("%Y-%m-%d %H:%M")
        end = dt.datetime(datenow.year, datenow.month, datenow.day, 23, 59)
        end = end + relativedelta(days=7)
        end = end.strftime("%Y-%m-%d %H:%M")
    elif dates == "Next Month":
        begin = dt.datetime(datenow.year, datenow.month, datenow.day, 0, 0)
        begin = begin.strftime("%Y-%m-%d %H:%M")
        end = dt.datetime(datenow.year, datenow.month, datenow.day, 23, 59)
        end = end + relativedelta(months = 1)
        end = end.strftime("%Y-%m-%d %H:%M")

    download(begin, end, internal, shared, window)

b = Button(mainframe, text="Ok", command = ok)
b.grid(row=2, column=2)

ws = root.winfo_screenwidth() # width of the screen
hs = root.winfo_screenheight() # height of the screen

x = (ws/2) - (350/2)
y = (hs/2) - (100/2)

root.geometry('%dx%d+%d+%d' % (350, 100, x, y))

root.mainloop()